"""XLSX parse — list_sheets + sheet-selective TabularData."""
from __future__ import annotations
import io
from typing import AsyncIterator, BinaryIO, Optional

from openpyxl import load_workbook

from filternarrange_engine.core.canonical import Column, TabularData
from filternarrange_engine.core.inference import infer_type_tag


def _read_workbook_bytes(source: BinaryIO) -> bytes:
    try:
        source.seek(0)
    except Exception:
        pass
    return source.read()


def list_sheets(source: BinaryIO) -> list[str]:
    blob = _read_workbook_bytes(source)
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def parse(source: BinaryIO, sheet_name: Optional[str] = None) -> TabularData:
    blob = _read_workbook_bytes(source)
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, ())
    header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]

    sample: list[tuple] = []
    for i, row in enumerate(rows_iter):
        if i >= 200:
            break
        sample.append(row)

    schema = [
        Column(
            name=name,
            type=infer_type_tag([row[i] for row in sample if i < len(row) and row[i] is not None]),
            nullable=True,
        )
        for i, name in enumerate(header)
    ]
    wb.close()

    wb2 = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws2 = wb2[sheet_name] if sheet_name else wb2[wb2.sheetnames[0]]
    it = ws2.iter_rows(values_only=True)
    next(it, None)
    all_dicts: list[dict] = []
    for row in it:
        all_dicts.append({header[i]: row[i] if i < len(row) else None for i in range(len(header))})
    wb2.close()

    return TabularData(schema=schema, rows=_aiter(all_dicts), meta={"row_count": len(all_dicts)})


async def _aiter(items) -> AsyncIterator[dict]:
    for r in items:
        yield r


__all__ = ["list_sheets", "parse"]
