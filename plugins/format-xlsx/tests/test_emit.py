import io
import pathlib
from openpyxl import load_workbook
from filternarrange_format_xlsx import plugin


def test_roundtrip():
    with open(pathlib.Path(__file__).parent / "fixtures/two_sheets.xlsx", "rb") as f:
        td = plugin.parse(f, sheet_name="people")
    sink = io.BytesIO()
    plugin.emit(td, sink)
    sink.seek(0)
    wb = load_workbook(sink, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(values_only=True))
    assert list(header) == ["id", "name", "age"]
    wb.close()
